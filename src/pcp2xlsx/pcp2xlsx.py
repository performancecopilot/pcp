#!/usr/bin/env pmpython
#
# Copyright (C) 2015-2018 Marko Myllynen <myllynen@redhat.com>
#
# This program is free software; you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by the
# Free Software Foundation; either version 2 of the License, or (at your
# option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of MERCHANTABILITY
# or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU General Public License
# for more details.

# pylint: disable=superfluous-parens
# pylint: disable=invalid-name, line-too-long, no-self-use
# pylint: disable=too-many-boolean-expressions, too-many-statements
# pylint: disable=too-many-instance-attributes, too-many-locals
# pylint: disable=too-many-branches, too-many-nested-blocks
# pylint: disable=broad-except

""" PCP to XLSX Bridge """

# Common imports
from collections import OrderedDict
import errno
import time
import sys

# Our imports
import os
import openpyxl

# PCP Python PMAPI
from pcp import pmapi, pmconfig
from cpmapi import PM_CONTEXT_ARCHIVE, PM_ERR_EOL, PM_IN_NULL, PM_DEBUG_APPL1

if sys.version_info[0] >= 3:
    long = int # pylint: disable=redefined-builtin

# Default config
DEFAULT_CONFIG = ["./pcp2xlsx.conf", "$HOME/.pcp2xlsx.conf", "$HOME/.pcp/pcp2xlsx.conf", "$PCP_SYSCONF_DIR/pcp2xlsx.conf"]

# Defaults
COLUMNA = 65
CONFVER = 1
PADDING = 5
TIMEFMT = "yyyy-mm-dd hh:mm:ss"

class PCP2XLSX(object):
    """ PCP to XLSX """
    def __init__(self):
        """ Construct object, prepare for command line handling """
        self.context = None
        self.daemonize = 0
        self.pmconfig = pmconfig.pmConfig(self)
        self.opts = self.options()

        # Configuration directives
        self.keys = ('source', 'output', 'derived', 'header', 'globals',
                     'samples', 'interval', 'type', 'precision', 'daemonize',
                     'timefmt',
                     'count_scale', 'space_scale', 'time_scale', 'version',
                     'count_scale_force', 'space_scale_force', 'time_scale_force',
                     'type_prefer', 'precision_force',
                     'speclocal', 'instances', 'ignore_incompat', 'omit_flat')

        # The order of preference for options (as present):
        # 1 - command line options
        # 2 - options from configuration file(s)
        # 3 - built-in defaults defined below
        self.check = 0
        self.version = CONFVER
        self.source = "local:"
        self.output = None # For pmrep conf file compat only
        self.speclocal = None
        self.derived = None
        self.header = 1
        self.globals = 1
        self.samples = None # forever
        self.interval = pmapi.timeval(60)      # 60 sec
        self.opts.pmSetOptionInterval(str(60)) # 60 sec
        self.delay = 0
        self.type = 0
        self.type_prefer = self.type
        self.ignore_incompat = 0
        self.instances = []
        self.omit_flat = 0
        self.precision = 3 # .3f
        self.precision_force = None
        self.timefmt = TIMEFMT
        self.interpol = 0
        self.count_scale = None
        self.count_scale_force = None
        self.space_scale = None
        self.space_scale_force = None
        self.time_scale = None
        self.time_scale_force = None

        # Not in pcp2xlsx.conf, won't overwrite
        self.outfile = None

        # Internal
        self.runtime = -1

        self.wb = None
        self.ws = None
        self.row = 0
        self.int_style = None
        self.float_style = None
        self.time_style = None

        # Performance metrics store
        # key - metric name
        # values - 0:txt label, 1:instance(s), 2:unit/scale, 3:type,
        #          4:width, 5:pmfg item, 6:precision, 7:limit
        self.metrics = OrderedDict()
        self.pmfg = None
        self.pmfg_ts = None

        # Read configuration and prepare to connect
        self.config = self.pmconfig.set_config_file(DEFAULT_CONFIG)
        self.pmconfig.read_options()
        self.pmconfig.read_cmd_line()
        self.pmconfig.prepare_metrics()
        self.pmconfig.set_signal_handler()

    def options(self):
        """ Setup default command line argument option handling """
        opts = pmapi.pmOptions()
        opts.pmSetOptionCallback(self.option)
        opts.pmSetOverrideCallback(self.option_override)
        opts.pmSetShortOptions("a:h:LK:c:Ce:D:V?HGA:S:T:O:s:t:rRIi:vP:0:q:b:y:Q:B:Y:F:f:Z:z")
        opts.pmSetShortUsage("[option...] metricspec [...]")

        opts.pmSetLongOptionHeader("General options")
        opts.pmSetLongOptionArchive()      # -a/--archive
        opts.pmSetLongOptionArchiveFolio() # --archive-folio
        opts.pmSetLongOptionContainer()    # --container
        opts.pmSetLongOptionHost()         # -h/--host
        opts.pmSetLongOptionLocalPMDA()    # -L/--local-PMDA
        opts.pmSetLongOptionSpecLocal()    # -K/--spec-local
        opts.pmSetLongOption("config", 1, "c", "FILE", "config file path")
        opts.pmSetLongOption("check", 0, "C", "", "check config and metrics and exit")
        opts.pmSetLongOption("output-file", 1, "F", "OUTFILE", "output file")
        opts.pmSetLongOption("derived", 1, "e", "FILE|DFNT", "derived metrics definitions")
        opts.pmSetLongOption("daemonize", 0, "", "", "daemonize on startup")
        opts.pmSetLongOptionDebug()        # -D/--debug
        opts.pmSetLongOptionVersion()      # -V/--version
        opts.pmSetLongOptionHelp()         # -?/--help

        opts.pmSetLongOptionHeader("Reporting options")
        opts.pmSetLongOption("no-header", 0, "H", "", "omit headers")
        opts.pmSetLongOption("no-globals", 0, "G", "", "omit global metrics")
        opts.pmSetLongOptionAlign()        # -A/--align
        opts.pmSetLongOptionStart()        # -S/--start
        opts.pmSetLongOptionFinish()       # -T/--finish
        opts.pmSetLongOptionOrigin()       # -O/--origin
        opts.pmSetLongOptionSamples()      # -s/--samples
        opts.pmSetLongOptionInterval()     # -t/--interval
        opts.pmSetLongOptionTimeZone()     # -Z/--timezone
        opts.pmSetLongOptionHostZone()     # -z/--hostzone
        opts.pmSetLongOption("raw", 0, "r", "", "output raw counter values (no rate conversion)")
        opts.pmSetLongOption("raw-prefer", 0, "R", "", "prefer output raw counter values (no rate conversion)")
        opts.pmSetLongOption("ignore-incompat", 0, "I", "", "ignore incompatible instances (default: abort)")
        opts.pmSetLongOption("instances", 1, "i", "STR", "instances to report (default: all current)")
        opts.pmSetLongOption("omit-flat", 0, "v", "", "omit single-valued metrics")
        opts.pmSetLongOption("precision", 1, "P", "N", "prefer N digits after decimal separator (default: 3)")
        opts.pmSetLongOption("precision-force", 1, "0", "N", "force N digits after decimal separator")
        opts.pmSetLongOption("timestamp-format", 1, "f", "STR", "xlsx timestamp format")
        opts.pmSetLongOption("count-scale", 1, "q", "SCALE", "default count unit")
        opts.pmSetLongOption("count-scale-force", 1, "Q", "SCALE", "forced count unit")
        opts.pmSetLongOption("space-scale", 1, "b", "SCALE", "default space unit")
        opts.pmSetLongOption("space-scale-force", 1, "B", "SCALE", "forced space unit")
        opts.pmSetLongOption("time-scale", 1, "y", "SCALE", "default time unit")
        opts.pmSetLongOption("time-scale-force", 1, "Y", "SCALE", "forced time unit")

        return opts

    def option_override(self, opt):
        """ Override standard PCP options """
        if opt in ('g', 'H', 'K', 'n', 'N', 'p'):
            return 1
        return 0

    def option(self, opt, optarg, index):
        """ Perform setup for an individual command line option """
        if opt == 'daemonize':
            self.daemonize = 1
        elif opt == 'K':
            if not self.speclocal or not self.speclocal.startswith(";"):
                self.speclocal = ";" + optarg
            else:
                self.speclocal = self.speclocal + ";" + optarg
        elif opt == 'c':
            self.config = optarg
        elif opt == 'C':
            self.check = 1
        elif opt == 'F':
            if os.path.exists(optarg):
                sys.stderr.write("File %s already exists.\n" % optarg)
                sys.exit(1)
            self.outfile = optarg
        elif opt == 'e':
            if not self.derived or not self.derived.startswith(";"):
                self.derived = ";" + optarg
            else:
                self.derived = self.derived + ";" + optarg
        elif opt == 'H':
            self.header = 0
        elif opt == 'G':
            self.globals = 0
        elif opt == 'r':
            self.type = 1
        elif opt == 'R':
            self.type_prefer = 1
        elif opt == 'I':
            self.ignore_incompat = 1
        elif opt == 'i':
            self.instances = self.instances + self.pmconfig.parse_instances(optarg)
        elif opt == 'v':
            self.omit_flat = 1
        elif opt == 'P':
            self.precision = optarg
        elif opt == '0':
            self.precision_force = optarg
        elif opt == 'f':
            self.timefmt = optarg
        elif opt == 'q':
            self.count_scale = optarg
        elif opt == 'Q':
            self.count_scale_force = optarg
        elif opt == 'b':
            self.space_scale = optarg
        elif opt == 'B':
            self.space_scale_force = optarg
        elif opt == 'y':
            self.time_scale = optarg
        elif opt == 'Y':
            self.time_scale_force = optarg
        else:
            raise pmapi.pmUsageErr()

    def connect(self):
        """ Establish a PMAPI context """
        context, self.source = pmapi.pmContext.set_connect_options(self.opts, self.source, self.speclocal)

        self.pmfg = pmapi.fetchgroup(context, self.source)
        self.pmfg_ts = self.pmfg.extend_timestamp()
        self.context = self.pmfg.get_context()

        if pmapi.c_api.pmSetContextOptions(self.context.ctx, self.opts.mode, self.opts.delta):
            raise pmapi.pmUsageErr()

    def validate_config(self):
        """ Validate configuration options """
        if self.version != CONFVER:
            sys.stderr.write("Incompatible configuration file version (read v%s, need v%d).\n" % (self.version, CONFVER))
            sys.exit(1)

        self.pmconfig.validate_common_options()

        if not self.outfile:
            sys.stderr.write("Output file must be defined.\n")
            sys.exit(1)

        self.time_style = openpyxl.styles.NamedStyle(name="datetime", number_format=self.timefmt)

        self.pmconfig.validate_metrics(curr_insts=True)
        self.pmconfig.finalize_options()

    def execute(self):
        """ Fetch and report """
        # Debug
        if self.context.pmDebug(PM_DEBUG_APPL1):
            sys.stdout.write("Known config file keywords: " + str(self.keys) + "\n")
            sys.stdout.write("Known metric spec keywords: " + str(self.pmconfig.metricspec) + "\n")

        # Set delay mode, interpolation
        if self.context.type != PM_CONTEXT_ARCHIVE:
            self.delay = 1
            self.interpol = 1

        # Common preparations
        self.context.prepare_execute(self.opts, False, self.interpol, self.interval)

        # Headers
        if self.header == 1:
            self.header = 0
            self.write_header()

        # Just checking
        if self.check == 1:
            return

        # Daemonize when requested
        if self.daemonize == 1:
            self.opts.daemonize()

        # Align poll interval to host clock
        if self.context.type != PM_CONTEXT_ARCHIVE and self.opts.pmGetOptionAlignment():
            align = float(self.opts.pmGetOptionAlignment()) - (time.time() % float(self.opts.pmGetOptionAlignment()))
            time.sleep(align)

        # Main loop
        while self.samples != 0:
            # Fetch values
            try:
                self.pmfg.fetch()
            except pmapi.pmErr as error:
                if error.args[0] == PM_ERR_EOL:
                    break
                raise error

            # Watch for endtime in uninterpolated mode
            if not self.interpol:
                if float(self.pmfg_ts().strftime('%s')) > float(self.opts.pmGetOptionFinish()):
                    break

            # Report and prepare for the next round
            self.report(self.pmfg_ts())
            if self.samples and self.samples > 0:
                self.samples -= 1
            if self.delay and self.interpol and self.samples != 0:
                self.pmconfig.pause()

        # Allow to flush buffered values / say goodbye
        self.report(None)

    def report(self, tstamp):
        """ Report the metric values """
        if tstamp != None:
            tstamp = tstamp.strftime(self.timefmt)

        self.write_xlsx(tstamp)

    def write_header(self):
        """ Write info header """
        if self.context.type == PM_CONTEXT_ARCHIVE:
            sys.stdout.write("Recording %d archived metrics to %s...\n(Ctrl-C to stop)\n" % (len(self.metrics), self.outfile))
            return

        sys.stdout.write("Recording %d metrics to %s" % (len(self.metrics), self.outfile))
        if self.runtime != -1:
            sys.stdout.write(":\n%s samples(s) with %.1f sec interval ~ %d sec runtime.\n" % (self.samples, float(self.interval), self.runtime))
        elif self.samples:
            duration = (self.samples - 1) * float(self.interval)
            sys.stdout.write(":\n%s samples(s) with %.1f sec interval ~ %d sec runtime.\n" % (self.samples, float(self.interval), duration))
        else:
            sys.stdout.write("...\n(Ctrl-C to stop)\n")

    def write_xlsx(self, timestamp):
        """ Write results in XLSX format """
        if timestamp is None:
            # Complete and close
            self.wb.save(self.outfile)
            self.wb.close()
            self.wb = None
            return

        # Current row
        self.row += 1

        # Alignments
        left = openpyxl.styles.Alignment(horizontal="left")
        right = openpyxl.styles.Alignment(horizontal="right")

        def cell_str(col, row):
            """ Helper to return cell string """
            return chr(col) + str(row)

        def write_cell(col, row, value=None, bold=False, align=right):
            """ Write value to cell """
            if value is None:
                self.ws.cell(col, self.row)
                return
            if bold:
                self.ws[cell_str(col, row)].font = openpyxl.styles.Font(bold=True)
            self.ws[cell_str(col, self.row)].alignment = align
            self.ws[cell_str(col, self.row)] = value

        # Create workbook, worksheet, and write headers
        if not self.wb:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "PCP Metrics"
            l = len(self.timefmt) if len(self.timefmt) > len("Timezone") else len("Timezone")
            self.ws.column_dimensions[chr(COLUMNA)].width = l + PADDING

            col = COLUMNA
            write_cell(col, self.row, "Host", True, left)
            col += 1
            write_cell(col, self.row, self.context.pmGetContextHostName(), True, left)
            l = len(self.context.pmGetContextHostName())
            self.ws.column_dimensions[chr(col)].width = l + PADDING
            self.row += 1

            col = COLUMNA
            write_cell(col, self.row, "Source", True, left)
            col += 1
            write_cell(col, self.row, self.source, True, left)
            l = len(self.source) if len(self.source) > l else l
            self.ws.column_dimensions[chr(col)].width = l + PADDING
            self.row += 1

            col = COLUMNA
            write_cell(col, self.row, "Timezone", True, left)
            col += 1
            timez = self.context.posix_tz_to_utc_offset(self.context.get_current_tz(self.opts))
            write_cell(col, self.row, timez, True, left)
            l = len(timez) if len(timez) > l else l
            self.ws.column_dimensions[chr(col)].width = l + PADDING
            self.row += 1

            # Add empty line for readability
            col = COLUMNA
            write_cell(col, self.row)
            self.row += 1

            col = COLUMNA
            write_cell(col, self.row, "Time", True)

            # Metrics/instances, static
            for i, metric in enumerate(self.metrics):
                for j in range(len(self.pmconfig.insts[i][0])):
                    col += 1
                    key = metric
                    if self.pmconfig.insts[i][0][0] != PM_IN_NULL and  self.pmconfig.insts[i][1][j]:
                        key += "[" + self.pmconfig.insts[i][1][j] + "]"
                    # Mark metrics with instance domain but without instances
                    if self.pmconfig.descs[i].contents.indom != PM_IN_NULL and self.pmconfig.insts[i][1][0] is None:
                        key += "[]"
                    write_cell(col, self.row, key, True)
                    l = len(key) if self.metrics[metric][4] < len(key) else self.metrics[metric][4]
                    if self.ws.column_dimensions[chr(col)].width is None or \
                       self.ws.column_dimensions[chr(col)].width < l + PADDING:
                        self.ws.column_dimensions[chr(col)].width = l + PADDING
            self.row += 1

            # Units, static
            col = COLUMNA
            for i, metric in enumerate(self.metrics):
                unit = self.metrics[metric][2][0]
                ins = 1 if self.pmconfig.insts[i][0][0] == PM_IN_NULL else len(self.pmconfig.insts[i][0])
                for _ in range(ins):
                    col += 1
                    write_cell(col, self.row, unit, True)
            self.row += 1

            # Add empty line with border for readability
            col = COLUMNA
            border = openpyxl.styles.borders.Border(top=openpyxl.styles.borders.Side(style="medium"))
            self.ws[cell_str(col, self.row)].border = border
            for i in range(len(self.metrics)):
                for _ in range(len(self.pmconfig.insts[i][0])):
                    col += 1
                    self.ws[cell_str(col, self.row)].border = border
            self.row += 1

            # Set number styles
            self.int_style = openpyxl.styles.NamedStyle(name="integer", number_format="0")
            precision = "0" if not self.metrics[metric][6] else "0." + "0" * self.metrics[metric][6]
            self.float_style = openpyxl.styles.NamedStyle(name="floating", number_format=precision)

        results = self.pmconfig.get_sorted_results()

        res = {}
        for i, metric in enumerate(results):
            for inst, _, value in results[metric]:
                res[metric + "+" + str(inst)] = value

        # Add corresponding values for each column in the static header
        col = COLUMNA
        self.ws[cell_str(col, self.row)].style = self.time_style
        self.ws[cell_str(col, self.row)] = self.pmfg_ts()
        for i, metric in enumerate(self.metrics):
            for j in range(len(self.pmconfig.insts[i][0])):
                col += 1
                try:
                    value = res[metric + "+" + str(self.pmconfig.insts[i][0][j])]
                    if value is None:
                        write_cell(col, self.row)
                    elif isinstance(value, str):
                        write_cell(col, self.row, value)
                    elif isinstance(value, float):
                        value = round(value, self.metrics[metric][6])
                        self.ws[cell_str(col, self.row)].style = self.float_style
                        write_cell(col, self.row, value)
                    else:
                        self.ws[cell_str(col, self.row)].style = self.int_style
                        write_cell(col, self.row, value)
                except Exception:
                    write_cell(col, self.row)

    def finalize(self):
        """ Finalize and clean up """
        if self.wb:
            self.wb.save(self.outfile)
            self.wb.close()
            self.wb = None
        return

if __name__ == '__main__':
    try:
        P = PCP2XLSX()
        P.connect()
        P.validate_config()
        P.execute()
        P.finalize()

    except pmapi.pmErr as error:
        sys.stderr.write('%s: %s\n' % (error.progname(), error.message()))
        sys.exit(1)
    except pmapi.pmUsageErr as usage:
        usage.message()
        sys.exit(1)
    except IOError as error:
        if error.errno != errno.EPIPE:
            sys.stderr.write("%s\n" % str(error))
            sys.exit(1)
    except KeyboardInterrupt:
        sys.stdout.write("\n")
        P.finalize()
